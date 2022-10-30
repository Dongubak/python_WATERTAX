import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
# openpyxl.load_workbook(파일명) : 파일 명을 인수로 하여 워크북의 객체로 반환한다.
# 해당하는 파일이 없는 경우 에러가 발생한다.

print(wb.sheetnames)
# workbook_object.sheetnames : 워크북에 존재하는 시트명을 리스트 형태로 취득할 수 있다.
# 원하는 시트에 있는 데이터만을 취득하고 싶다면 workbook_object[시트명]로 얻을 수 있다.
# 시트의 데이터를 얻으면 셀의 값을 다룰 수 있다.

# 셀의 값 구하기

# workbook_object[셀].value
sheet1 = wb["test_sheet_1"]
print(sheet1["A1"].value)

# cell
sheet1 = wb["test_sheet_1"]
print(sheet1.cell(1, 1).value)

# 시트 명 변경하기
# workbook_object.active : 워크북에서 active상태인 시트 객체 데이터를 얻을 수 있다.
sheet = wb.active
# title 속성을 통해 시트명 변경 가능
sheet.title = 'test_sheet_1'
wb.save('example.xlsx')


# 새로운 excel파일 만들기
# openpyxl.Workbook() : 내부가 비어있는 새로운 워크북 객체를 생성한다.
# save메서드를 이용하여 저장을 해야지만 엑셀 파일이 생성되었다고 할 수 있다.
wb = openpyxl.Workbook()
# wb.save('test.xlsx')

# 시트 추가
wb.create_sheet()
# 별다른 규칙 없이 Sheet숫자가 이름이 된다.
# Sheet는 openpyxl.Workbook()함수를 실행시키면 자동으로 생성되는 시트이다.

#wb.create_sheet(index = 숫자, title = 시트이름)
wb.create_sheet(index=0, title="새로운 시트")
wb.save('test.xlsx')

# 시트 삭제
# workbook_object.remove_sheet(시트객체)
# wb.remove_sheet(wb['Sheet1'])

# 시트 복제
# copy가 된 sheet의 이름은 시트객체명 Copy가 된다.
# copy하고 그 copy된 sheet의 title을 바꾸고 save를 합니다.
wb = openpyxl.load_workbook('example.xlsx')
wb.copy_worksheet(wb['test_sheet_1'])
copy_sheet = wb['test_sheet_1 Copy']
copy_sheet.title = 'Sheet7'
wb.save('example.xlsx')

# Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'test'
wb.save('test_write.xlsx')
