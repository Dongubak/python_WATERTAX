# 1. 수도요금 조사표 작성하기
# 가장 최근의 수도요금 공지에서 당월 값을 구한다.
# 구한 당원값들을 수도요금 조사 sheet의 B3부터 B11에 적용

import openpyxl
from openpyxl import Workbook


def convert_month(last_month: str) -> str:
    # input lastmonth is type of string, so convert string to int
    last_month: int = int(last_month)

    next_month = 1 if last_month == 12 else last_month + 1
    return str(next_month)


def get_invest_water() -> None:
    wb = openpyxl.load_workbook('수도요금(최신버전).xlsx')
    latest_sheet_name = wb.sheetnames[-1]
    invest_water = wb.sheetnames[0]
    latest_sheet = wb[latest_sheet_name]
    invest_sheet = wb[invest_water]
    prev_month_water = []
    for i in range(10):
        if(i == 2):
            continue
        prev_month_water.append(latest_sheet[f'C{i + 2}'].value)
    index = 3
    for i in prev_month_water:
        invest_sheet[f'B{index}'].value = i
        index = index + 1
    wb.save('수도요금(최신버전).xlsx')


# 2. 수도요금 계산하기
# 가장 최근의 수도요금 공지를 복사한다.
# 복사한 수도요금 공지의 sheet명을 다음 월로 변경한다.
# 사용자의 당월 값 입력을 받아 복사한 sheet에서 수도요금 계산을 진행한다.
# 이때 새로 생성된 sheet는 오른쪽에 위치한다.
def cal_water_tax():
    wb = openpyxl.load_workbook('수도요금(최신버전).xlsx')
    latest_sheet_name = wb.sheetnames[-1]
    latest_sheet = wb[latest_sheet_name]
    wb.copy_worksheet(latest_sheet)
    copy_sheet = wb[latest_sheet_name + ' Copy']
    new_sheet_name = convert_month(latest_sheet_name[:2])
    copy_sheet.title = f'{new_sheet_name}월 수도요금 공지'
    print('--------------------------------------------------------------------')
    print('수도요금 자동화 계산기(엑셀자동업데이트)')
    print('--------------------------------------------------------------------')
    print('B01 B02 102 201 202 301 302 401호의 수도사용량을 입력하시오')
    homes = ['B01', 'B02', '101', '102', '201',
             '202', '301', '302', '401', '402']
    input_water = []

    for i in range(10):
        if i == 2:
            continue
        copy_sheet[f'B{i + 2}'].value = copy_sheet[f'C{i + 2}'].value

    for i in homes:
        if i == '101':
            continue
        input_data = input(f'{i} : ')
        input_water.append(input_data)
    print(input_water)
    for i in range(9):
        if i < 2:
            copy_sheet[f'C{i + 2}'].value = input_water[i]
        else:
            copy_sheet[f'C{i + 3}'].value = input_water[i]
    print('--------------------------------------------------------------------')
    water_price = input('상하수도 요금을 입력하시오 : ')
    copy_sheet['D13'] = water_price

    wb.save('수도요금(최신버전).xlsx')


cal_water_tax()
